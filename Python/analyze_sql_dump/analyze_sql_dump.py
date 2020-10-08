import os, re, openpyxl
from openpyxl.styles.borders import Border, Side, BORDER_MEDIUM, BORDER_THICK
from openpyxl.styles import PatternFill, Color, Font

med_border = Border(
    left=Side(border_style=BORDER_MEDIUM, color='00000000'),
    right=Side(border_style=BORDER_MEDIUM, color='00000000'),
    top=Side(border_style=BORDER_MEDIUM, color='00000000'),
    bottom=Side(border_style=BORDER_MEDIUM, color='00000000')
)
fat_border = Border(
    left=Side(border_style=BORDER_THICK, color='00000000'),
    right=Side(border_style=BORDER_THICK, color='00000000'),
    top=Side(border_style=BORDER_THICK, color='00000000'),
    bottom=Side(border_style=BORDER_THICK, color='00000000')
)
black_background = PatternFill(start_color='00000000',
                   end_color='00000000',
                   fill_type='solid')

grey_background = PatternFill(start_color='AAAAAAAA',
                   end_color='AAAAAAAA',
                   fill_type='solid')



def analyze_data(file_content):
    pos_tuples = []
    startpoint = None
    # Find table definitions
    for i in range(0,len(file_content)):
        # Start of table
        if "CREATE TABLE" in file_content[i] and startpoint == None:
            startpoint = i
        # End of table
        elif "ENGINE" in file_content[i] and startpoint != None:
            pos_tuples.append((startpoint, i))
            startpoint = None

    table_definitions = {}
    # Regex patterns
    table_name_pattern = re.compile(r'`(.*)`')
    datatype_pattern = re.compile(r'` ([a-zA-Z0-9)(]+)')
    defaut_pattern = re.compile(r'DEFAULT (.*),?')
    foreign_pattern = re.compile(r'.*FOREIGN KEY \((.*)\) REFERENCES (.*)\((.*)\)')
    primary_key_pattern = re.compile(r'.*PRIMARY KEY \(`(.*)`\)')

    for position in pos_tuples:
        # Iterate trough table definition
        for i in range(position[0], position[1]+1):
            # Table name definition
            if i == position[0]:
                table_name = table_name_pattern.findall(file_content[i])[0]
                table_definitions.setdefault(table_name, {})
            # Read table attributes
            elif not "ENGINE" in file_content[i].upper() and not "PRIMARY" in file_content[i].upper() and not "FOREIGN" in file_content[i].upper() and not "/*" in file_content[i]: 
                try:
                    attr_name = table_name_pattern.findall(file_content[i])[0]
                    table_definitions[table_name].setdefault(attr_name, {})

                    datatype = datatype_pattern.findall(file_content[i])[0]
                    table_definitions[table_name][attr_name].setdefault("DATATYPE", datatype)

                    if "NOT NULL" in file_content[i].upper():
                        table_definitions[table_name][attr_name].setdefault("NULLABLE", False)
                    else:
                        table_definitions[table_name][attr_name].setdefault("NULLABLE", True)

                    if "AUTO_INCREMENT" in file_content[i].upper():
                        table_definitions[table_name][attr_name].setdefault("AUTO_INCREMENT", True)

                    if "DEFAULT " in file_content[i].upper():
                        default_value = defaut_pattern.findall(file_content[i].upper())[0]
                        default_value = default_value.strip(" ")
                        default_value = default_value.strip(",")
                        default_value = default_value.strip("'")

                        if default_value == "":
                            default_value = "NULL"

                        table_definitions[table_name][attr_name].setdefault("DEFAULT_VALUE", default_value)
                    else:
                        table_definitions[table_name][attr_name].setdefault("DEFAULT_VALUE", None)
                except:
                    pass

            # Get primary_key
            elif "PRIMARY" in file_content[i].upper():
                try:
                    primary_key = primary_key_pattern.findall(file_content[i].upper())[0]
                    table_definitions[table_name].setdefault("PRIMARY_KEY", primary_key)
                except:
                    pass
            # Get foreign_key
            elif "FOREIGN" in file_content[i].upper():
                try:
                    foreign_key_values = foreign_pattern.findall(file_content[i].upper())
                    keyname = foreign_key_values[0][0].strip("`")
                    target_table = foreign_key_values[0][1]
                    target_attribute = foreign_key_values[0][2].strip("`")
                    table_definitions[table_name].setdefault("FOREIGN_KEY", {})
                    table_definitions[table_name]["FOREIGN_KEY"].setdefault(keyname.lower(), { "TARGET_TABLE" : target_table, "TARGET_ATTR" : target_attribute})
                except:
                    pass
                

    return table_definitions

def choose_file():
    while True:
        path = input("Pfad zur SQL-Datei angeben: ")
        if path.startswith('"'):
            path = path.strip('"')
        if os.path.exists(path) and os.path.isfile(path):
            return path
        else:
            print("Datei {} konnte nicht gefunden werden".format(path))


def get_target_file_path():
    while True:
        path = input("Pfad zur Zieldatei angeben: ")
        if os.path.exists(path) and not os.path.isfile(path):
            if "/" in path:
                if path.endswith("/"):
                    path = path[:-1]
                delimiter = "/"
            elif "\\" in path:
                if path.endswith("\\"):
                    path = path[:-1]
                delimiter = "\\"
            break
        else:
            print("Pfad {} existiert nicht oder ist eine Datei.".format(path))
            
    while True:
        filename = input("Dateiname angeben: ")
        if not filename.endswith(".xlsx"):
            filename += ".xlsx"
        abs_path = path + delimiter + filename
        if os.path.exists(abs_path) and os.path.isfile(abs_path):
            overwrite = input("Die Datei {} existiert bereits. Überschreiben? [Y/n]: ")
            if not overwrite in [ "y", "Y", "J", "j", ""]:
                continue
            os.remove(abs_path)
        break
    
    return abs_path


def get_file_content(path):
    file_reader = open(path)
    file_content = file_reader.read()
    file_content = file_content.split("\n")
    return file_content


def output_data(dictionary, target_file):
    excel_sheet = openpyxl.Workbook()
    sheet_names = excel_sheet.sheetnames
    active_sheet = excel_sheet[sheet_names[0]]
    active_sheet.title = "Dump analysis"
    active_sheet.freeze_panes = "A2"



    table_template = { "A1" : "Tabellenname", 
                       "B1" : "Attributname", 
                       "C1" : "Datentyp", 
                       "D1" : "Nullbar", 
                       "E1" : "Standardwert", 
                       "F1" : "Extra", 
                       "G1" : "Schlüssel"
    }

    for cell in table_template.keys():
        active_sheet[cell] = table_template[cell]

    y_axis = 2

    tables = dictionary.keys()
    for table in tables:
        attributes = dictionary[table].keys()
        have_foreign_key = False
        if "PRIMARY_KEY" in attributes:
            primary_key_name = dictionary[table]["PRIMARY_KEY"]
            del dictionary[table]["PRIMARY_KEY"]
        if "FOREIGN_KEY" in dictionary[table].keys():
            have_foreign_key = True
            foreign_keys = dictionary[table]["FOREIGN_KEY"]
            local_foreign_key_names = dictionary[table]["FOREIGN_KEY"].keys()

            del dictionary[table]["FOREIGN_KEY"]
        
        for attribute in attributes:
            active_sheet["A"+str(y_axis)] = table
            active_sheet["B"+str(y_axis)] = attribute
            active_sheet["C"+str(y_axis)] = dictionary[table][attribute]["DATATYPE"]
            active_sheet["D"+str(y_axis)] = dictionary[table][attribute]["NULLABLE"]
            active_sheet["E"+str(y_axis)] = dictionary[table][attribute]["DEFAULT_VALUE"]
            if "AUTO_INCREMENT" in dictionary[table][attribute].keys():
                current_value = active_sheet["F"+str(y_axis)].value
                if current_value == None:
                    active_sheet["F"+str(y_axis)] = "auto_increment"
                else:
                    active_sheet["F"+str(y_axis)] = current_value + ", auto_increment"
            if attribute.lower() == primary_key_name.lower():
                active_sheet["G"+str(y_axis)] = "PRIMÄRSCHLÜSSEL"
            elif have_foreign_key and attribute in local_foreign_key_names:
                target_foreign_key_name = foreign_keys[attribute]["TARGET_ATTR"]
                target_foreign_key_table = foreign_keys[attribute]["TARGET_TABLE"]
                active_sheet["G"+str(y_axis)] = "Fremschlüssel nach {}".format(target_foreign_key_table + "." + target_foreign_key_name)
            y_axis += 1

    max_row = active_sheet.max_row
    max_col = active_sheet.max_column
    last_table = None
    invert_color = False

    for i in range(1,max_row+1):
        current_table = active_sheet.cell(row=i, column=1).value
        for j in range (1, max_col+1):
            if i == 1:
                active_sheet.cell(row=i, column=j).border = fat_border
                active_sheet.cell(row=i, column=j).fill = black_background
                active_sheet.cell(row=i, column=j).font = Font(color = "FFFFFF", bold = True)
            else:
                if last_table == None or last_table != current_table:
                    last_table = current_table
                    invert_color = not invert_color
                if invert_color:
                    active_sheet.cell(row=i, column=j).fill = grey_background
                active_sheet.cell(row=i, column=j).border = med_border
            if j == 1 and i != 1:
                active_sheet.cell(row=i, column=j).font = Font(bold = True)


    for column_cells in active_sheet.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        active_sheet.column_dimensions[column_cells[0].column_letter].width = length

    excel_sheet.save(target_file)


    # for key in dictionary.keys():
    #     print("\t"*depth + key.upper())
    #     if type(dictionary[key]) == dict:
    #         depth += 1
    #         output_data(dictionary[key], depth)
    #         depth -= 1
    #     else:
    #         print("\t"*depth + str(dictionary[key]).lower())

path = choose_file()
target_file = get_target_file_path()
content_list = get_file_content(path)
tables = analyze_data(content_list)
output_data(tables, target_file)

